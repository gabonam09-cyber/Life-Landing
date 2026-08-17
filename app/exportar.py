"""Descarga de los datos del panel en un archivo de Excel.

Sirve de dos cosas: tener los prospectos a mano para trabajarlos fuera del panel,
y guardar una copia propia de todo. La base de datos vive en un servicio ajeno y
puede caducar o fallar; un .xlsx en tu disco no depende de nadie.
"""

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .analitica import ETIQUETAS_SECCIONES, ORDEN_SECCIONES
from .models import Lead, Sesion

# Colores de la marca, para que el archivo no se vea como una tabla cualquiera.
VERDE = "CCFF00"
NEGRO = "1A1A1A"


def _encabezar(hoja, titulos):
    hoja.append(titulos)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color=NEGRO, size=11)
        celda.fill = PatternFill("solid", fgColor=VERDE)
        celda.alignment = Alignment(vertical="center")
    hoja.freeze_panes = "A2"  # los titulos quedan fijos al desplazarse


def _ajustar_anchos(hoja, maximo=60):
    """Cada columna se ensancha a su contenido, sin pasarse de largo."""
    for columna in hoja.columns:
        letra = get_column_letter(columna[0].column)
        largo = max((len(str(c.value)) for c in columna if c.value is not None), default=8)
        hoja.column_dimensions[letra].width = min(maximo, max(11, largo + 3))


def _fecha(valor):
    return valor.strftime("%Y-%m-%d %H:%M") if valor else ""


def _hoja_prospectos(libro):
    hoja = libro.active
    hoja.title = "Prospectos"
    _encabezar(hoja, ["Fecha", "Nombre", "Negocio", "Correo", "Teléfono", "Qué necesita"])

    for lead in Lead.query.order_by(Lead.creado_en.desc()).all():
        hoja.append([_fecha(lead.creado_en), lead.nombre, lead.clinica or "",
                     lead.email or "", lead.telefono or "", lead.mensaje or ""])

    # El mensaje suele ser largo: se envuelve en vez de desbordar la celda.
    for fila in hoja.iter_rows(min_row=2, min_col=6, max_col=6):
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, vertical="top")

    _ajustar_anchos(hoja)
    hoja.column_dimensions["F"].width = 55
    return hoja


def _hoja_visitas(libro, sesiones):
    hoja = libro.create_sheet("Visitas")
    _encabezar(hoja, ["Fecha", "Origen", "Dispositivo", "Duración (s)",
                      "Scroll máximo (%)", "Vio el video", "Segundos de video",
                      "% del video", "Clics", "Visita propia"])

    for s in sesiones:
        video = s.video
        hoja.append([
            _fecha(s.creada_en), s.origen, s.dispositivo,
            s.duracion_seg or 0, s.scroll_max_pct or 0,
            "Sí" if (video and video.dio_play) else "No",
            (video.segundos_vistos or 0) if video else 0,
            (video.pct_completado or 0) if video else 0,
            len(s.clics),
            "Sí" if s.es_interna else "No",
        ])

    _ajustar_anchos(hoja)


def _hoja_secciones(libro, sesiones):
    """El embudo: hasta donde llega la gente y cuanto aguanta en cada bloque."""
    hoja = libro.create_sheet("Recorrido")
    _encabezar(hoja, ["Sección", "Visitas que llegaron", "% del total",
                      "Segundos promedio"])

    externas = [s for s in sesiones if not s.es_interna]
    total = len(externas)

    acumulado = {}
    for s in externas:
        for sv in s.secciones:
            if sv.segundos > 0:
                d = acumulado.setdefault(sv.seccion, {"n": 0, "seg": 0})
                d["n"] += 1
                d["seg"] += sv.segundos

    def orden(nombre):
        return ORDEN_SECCIONES.index(nombre) if nombre in ORDEN_SECCIONES else 999

    for nombre, d in sorted(acumulado.items(), key=lambda x: orden(x[0])):
        hoja.append([
            ETIQUETAS_SECCIONES.get(nombre, nombre),
            d["n"],
            round(d["n"] * 100 / total) if total else 0,
            round(d["seg"] / d["n"]),
        ])

    _ajustar_anchos(hoja)


def _hoja_clics(libro, sesiones):
    hoja = libro.create_sheet("Clics")
    _encabezar(hoja, ["Fecha", "Dispositivo", "Elemento", "Texto",
                      "Posición horizontal (%)", "Altura en la página (px)"])

    for s in sesiones:
        for c in s.clics:
            hoja.append([_fecha(c.creado_en), c.dispositivo, c.elemento or "",
                         (c.texto or "").replace("\n", " ")[:80],
                         round(c.x_rel * 100), c.y_abs])

    _ajustar_anchos(hoja)


def construir_excel():
    """Devuelve el .xlsx completo en memoria, listo para descargar."""
    libro = Workbook()

    _hoja_prospectos(libro)

    sesiones = Sesion.query.order_by(Sesion.creada_en.desc()).all()
    _hoja_visitas(libro, sesiones)
    _hoja_secciones(libro, sesiones)
    _hoja_clics(libro, sesiones)

    memoria = BytesIO()
    libro.save(memoria)
    memoria.seek(0)
    return memoria


def nombre_archivo():
    return "life-datos-" + datetime.now(timezone.utc).strftime("%Y-%m-%d") + ".xlsx"
